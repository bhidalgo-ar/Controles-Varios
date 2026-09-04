# -*- coding: utf-8 -*-
"""El Excel que ve el analista. Cinco hojas, siempre las mismas, mas la de
contribuciones cuando vino el control de cargas sociales.

El criterio de lectura: el numero grande es la diferencia de neto, y para cada
legajo que no cierra tiene que estar escrito EN QUE CONCEPTO esta la plata.
Las diferencias de centavos no llevan comentario: ensucian el archivo y hacen
que el analista deje de leer la columna.
"""
import collections
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

AZUL = "1F3864"; ROJO = "C00000"; VERDE = "375623"; AMBAR = "7F6000"
MON = '#,##0.00'
H = Font(bold=True, color="FFFFFF", size=10)
HF = PatternFill("solid", fgColor=AZUL)
B = Font(bold=True, size=10)
N = Font(size=10)
BOX = Border(bottom=Side(style='thin', color="BFBFBF"))


def escribir(res, ancla, meta, salida, tol=0.01, ruido=1.00, contrib=None):
    wb = openpyxl.Workbook()
    porLegajo, detalle = res['porLegajo'], res['detalle']
    totM4 = sum(v['netoM4'] for v in porLegajo)
    totAx = sum(v['netoAx'] for v in porLegajo)
    conDif = [v for v in porLegajo if abs(v['difNeto']) > tol]
    conDifReal = [v for v in porLegajo if abs(v['difNeto']) > ruido]

    def cab(ws, cols, fila=1):
        for i, (t, w) in enumerate(cols, 1):
            c = ws.cell(row=fila, column=i, value=t)
            c.font = H; c.fill = HF
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = ws.cell(row=fila + 1, column=1)

    # ============================================== 1. Resumen
    ws = wb.active; ws.title = "Resumen"
    ws.column_dimensions['A'].width = 50
    for col in 'BCDEFGHI':
        ws.column_dimensions[col].width = 18
    est = {'r': 1}

    def tit(txt, size=13):
        ws.cell(row=est['r'], column=1, value=txt).font = Font(bold=True, size=size, color=AZUL)
        est['r'] += 2

    def kv(k, v, fmt=None, bold=False, color=None):
        ws.cell(row=est['r'], column=1, value=k).font = B if bold else N
        c = ws.cell(row=est['r'], column=2, value=v)
        c.font = Font(bold=bold, size=10, color=color or "000000") if (bold or color) else N
        if fmt:
            c.number_format = fmt
        est['r'] += 1

    def nota(txt, color=None):
        ws.cell(row=est['r'], column=1, value=txt).font = Font(size=9, italic=True,
                                                              color=color or "404040")
        est['r'] += 1

    tit(f"Paralelo {meta.get('cliente','')} — {meta.get('periodo','')}", 15)
    veredicto = ("CIERRA" if not conDifReal else
                 f"{len(conDifReal)} legajos con diferencia")
    kv("Estado", veredicto, bold=True, color=VERDE if not conDifReal else ROJO)
    kv("Lado tomado como correcto", "Tabulado de Meta4 + PDF de la liqui")
    kv("Lado controlado", "Tabulado de Axton")
    kv("Archivo de Axton", meta.get('axtonNombre', ''))
    kv("Encabezado del archivo de Axton", meta.get('preambulo', '')[:200])
    kv("Tipo de liquidación que trae", ", ".join(meta.get('tipos', []))[:200])
    kv("Equivalencias de concepto", f"{meta.get('pares', 0)} pares de la tabla")
    est['r'] += 1

    tit("Verificación previa: ¿estoy leyendo bien los archivos?", 12)
    kv("Neto del Excel de Meta4 (consolidado por legajo)", ancla['totalExcel'], MON)
    kv("Neto del PDF de la liqui (Total Netos)", ancla['totalPdf'], MON)
    kv("Diferencia entre los dos", round(ancla['totalExcel'] - ancla['totalPdf'], 2), MON,
       bold=True, color=VERDE if abs(ancla['totalExcel'] - ancla['totalPdf']) <= tol else ROJO)
    kv("Legajos donde el neto del Excel coincide con el PDF",
       f"{ancla['comparados'] - len(ancla['difieren']) - len(ancla['sinBloqueEnPdf'])} de {ancla['comparados']}")
    kv("Fichas de empleado que trae el PDF", ancla['bloquesPdf'])
    kv("Filas de Meta4 donde Haberes − Descuentos = NETO", meta.get('validaM4', ''))
    kv("Filas de Axton donde Bruto − Retenciones + Exento = Neto", meta.get('validaAx', ''))
    est['r'] += 1

    tit("El resultado", 12)
    kv("Empleados comparados", len(porLegajo))
    if res['soloM4']:
        kv("Legajos que están sólo en Meta4", ", ".join(res['soloM4'])[:200], color=ROJO)
    if res['soloAx']:
        kv("Legajos que están sólo en Axton", ", ".join(res['soloAx'])[:200], color=ROJO)
    kv("Liquidaciones en el archivo de Meta4", meta.get('filasM4', ''))
    kv("NETO total Meta4", round(totM4, 2), MON, bold=True)
    kv("NETO total Axton", round(totAx, 2), MON, bold=True)
    kv("Diferencia (Axton − Meta4)", round(totAx - totM4, 2), MON, bold=True,
       color=VERDE if abs(totAx - totM4) <= ruido else ROJO)
    kv("Legajos cuyo neto coincide exacto", len(porLegajo) - len(conDif), color=VERDE)
    kv("Legajos que difieren $1 o menos (redondeo)", len(conDif) - len(conDifReal), color=AMBAR)
    kv("Legajos con diferencia de verdad", len(conDifReal),
       color=ROJO if conDifReal else VERDE)
    sinExpl = [v for v in porLegajo if abs(v['sinExplicar']) > ruido]
    kv("Legajos cuya diferencia NO se explica con sus conceptos", len(sinExpl),
       color=ROJO if sinExpl else VERDE)
    if sinExpl:
        nota("Ojo: si este numero no es cero, hay un concepto que quedo afuera del mapeo "
             "y el detalle de esos legajos esta incompleto.", ROJO)
    est['r'] += 1

    tit("Dónde está la diferencia, concepto por concepto", 12)
    agg = collections.defaultdict(lambda: {'m4': 0.0, 'ax': 0.0, 'ap': 0.0,
                                           'tipo': '', 'cm': '', 'ca': '', 'legs': []})
    for f in detalle:
        d = agg[f['concepto']]
        d['tipo'], d['cm'], d['ca'] = f['tipo'], f['codM4'], f['codAx']
        d['m4'] += f['meta4'] or 0
        d['ax'] += f['axton'] or 0
        d['ap'] += f['aporteNeto']
        if abs(f['aporteNeto']) > ruido:
            d['legs'].append(f['legajo'])
    cols = [("Concepto", 44), ("Tipo", 12), ("Cód. Meta4", 12), ("Cód. Axton", 16),
            ("Total Meta4", 17), ("Total Axton", 17), ("Cuánto mueve el neto", 18),
            ("Legajos", 9), ("Cuáles", 36)]
    for i, (t, w) in enumerate(cols, 1):
        c = ws.cell(row=est['r'], column=i, value=t)
        c.font = H; c.fill = HF
        c.alignment = Alignment(horizontal='center', wrap_text=True)
        if i > 2:
            ws.column_dimensions[get_column_letter(i)].width = w
    est['r'] += 1
    hubo = False
    for k, d in sorted(agg.items(), key=lambda kv: -abs(kv[1]['ap'])):
        if abs(d['ap']) <= ruido:
            continue
        hubo = True
        for i, val in ((1, k), (2, d['tipo']), (3, d['cm']), (4, d['ca']),
                       (8, len(d['legs'])),
                       (9, ", ".join(d['legs'][:8]) + (" …" if len(d['legs']) > 8 else ""))):
            ws.cell(row=est['r'], column=i, value=val).font = N
        for i, val in ((5, d['m4']), (6, d['ax']), (7, d['ap'])):
            c = ws.cell(row=est['r'], column=i, value=round(val, 2))
            c.number_format = MON; c.font = N
        ws.cell(row=est['r'], column=7).font = Font(size=10, bold=True,
                                                   color=ROJO if d['ap'] < 0 else VERDE)
        est['r'] += 1
    if not hubo:
        ws.cell(row=est['r'], column=1,
                value="No hay ningún concepto que mueva el neto más de $1.").font = Font(
            size=10, italic=True, color=VERDE)
        est['r'] += 1
    est['r'] += 1
    nota("\"Cuánto mueve el neto\": cuánto sube (+) o baja (−) el neto de Axton por ese "
         "concepto. Un haber de menos y un descuento de más bajan el neto los dos.")
    nota(f"Las diferencias de ${ruido:,.2f} o menos no se comentan: son centavos de "
         "redondeo, no un hallazgo.")

    if meta.get('mapeosDeclarados'):
        est['r'] += 1
        tit("Emparejamientos que se declararon a mano (confirmar y corregir la tabla)", 12)
        for ax, extras in meta['mapeosDeclarados'].items():
            kv(f"Axton {ax}", "también se compara contra " + ", ".join(extras))

    # ============================================== 2. Netos por legajo
    ws = wb.create_sheet("Netos por legajo")
    cab(ws, [("Legajo", 10), ("Apellido y Nombre", 32), ("Liq. Meta4", 11),
             ("Neto Meta4", 16), ("Neto Axton", 16), ("Diferencia", 16),
             ("Estado", 20), ("Dónde está la diferencia", 110)])
    rr = 2
    for v in sorted(porLegajo, key=lambda v: -abs(v['difNeto'])):
        grande = abs(v['difNeto']) > ruido
        txt = ""
        if grande:
            difs = sorted([f for f in v['conceptos'] if abs(f['aporteNeto']) > ruido],
                          key=lambda f: -abs(f['aporteNeto']))
            txt = " | ".join(f"{f['concepto']} ({f['codM4']}→{f['codAx']}): "
                             f"{f['aporteNeto']:+,.2f}" for f in difs)
            ex = ([f"Meta4 {c} sin equivalencia: {x:,.2f}" for c, x in v['extraM4'] if abs(x) > ruido]
                  + [f"Axton {c} sin equivalencia: {x:,.2f}" for c, x in v['extraAx'] if abs(x) > ruido])
            if ex:
                txt = (txt + " | " if txt else "") + " | ".join(ex)
        etiqueta = ("Difiere" if grande else
                    ("Coincide" if abs(v['difNeto']) <= tol else "Difiere $1 o menos"))
        for i, val in enumerate([v['legajo'], v['nombre'], v['liqsM4'], v['netoM4'],
                                 v['netoAx'], v['difNeto'], etiqueta, txt], 1):
            c = ws.cell(row=rr, column=i, value=val)
            c.font = N; c.border = BOX
            if i in (4, 5, 6):
                c.number_format = MON
        if etiqueta == "Difiere":
            ws.cell(row=rr, column=6).font = Font(size=10, bold=True, color=ROJO)
            ws.cell(row=rr, column=7).font = Font(size=10, bold=True, color=ROJO)
        elif etiqueta == "Coincide":
            ws.cell(row=rr, column=7).font = Font(size=10, color=VERDE)
        else:
            ws.cell(row=rr, column=7).font = Font(size=10, color=AMBAR)
        rr += 1
    ws.auto_filter.ref = f"A1:H{rr-1}"
    ws.cell(row=rr + 1, column=3, value="TOTAL").font = B
    for i, val in ((4, totM4), (5, totAx), (6, round(totAx - totM4, 2))):
        c = ws.cell(row=rr + 1, column=i, value=round(val, 2))
        c.number_format = MON
        c.font = Font(bold=True, color=ROJO if i == 6 and abs(totAx - totM4) > ruido else "000000")

    # ============================================== 3. Diferencias por concepto
    ws = wb.create_sheet("Diferencias por concepto")
    cab(ws, [("Legajo", 10), ("Apellido y Nombre", 30), ("Concepto", 40), ("Tipo", 12),
             ("Cód. Meta4", 12), ("Cód. Axton", 16), ("Importe Meta4", 16),
             ("Importe Axton", 16), ("Axton − Meta4", 16), ("Mueve el neto", 16),
             ("Qué pasa", 32)])
    nombre = {v['legajo']: v['nombre'] for v in porLegajo}
    rr = 2
    for f in sorted(detalle, key=lambda f: -abs(f['aporteNeto'])):
        if abs(f['aporteNeto']) <= ruido:
            continue
        for i, val in enumerate([f['legajo'], nombre.get(f['legajo'], ''), f['concepto'],
                                 f['tipo'], f['codM4'], f['codAx'], f['meta4'], f['axton'],
                                 f['dif'], round(f['aporteNeto'], 2), f['estado']], 1):
            c = ws.cell(row=rr, column=i, value=val)
            c.font = N; c.border = BOX
            if i in (7, 8, 9, 10):
                c.number_format = MON
        ws.cell(row=rr, column=10).font = Font(size=10, bold=True,
                                              color=ROJO if f['aporteNeto'] < 0 else VERDE)
        if f['estado'].startswith('Falta'):
            ws.cell(row=rr, column=11).font = Font(size=10, color=AMBAR)
        rr += 1
    ws.auto_filter.ref = f"A1:K{rr-1}"

    # ============================================== 4. Detalle completo
    ws = wb.create_sheet("Detalle completo")
    cab(ws, [("Legajo", 10), ("Concepto", 40), ("Tipo", 12), ("Cód. Meta4", 12),
             ("Cód. Axton", 16), ("Importe Meta4", 16), ("Importe Axton", 16),
             ("Axton − Meta4", 16), ("Estado", 30)])
    rr = 2
    for f in sorted(detalle, key=lambda f: (int(f['legajo']) if f['legajo'].isdigit() else 0,
                                            f['concepto'])):
        if f['meta4'] in (None, 0) and f['axton'] in (None, 0):
            continue
        for i, val in enumerate([f['legajo'], f['concepto'], f['tipo'], f['codM4'],
                                 f['codAx'], f['meta4'], f['axton'], f['dif'],
                                 f['estado']], 1):
            c = ws.cell(row=rr, column=i, value=val)
            c.font = N
            if i in (6, 7, 8):
                c.number_format = MON
        rr += 1
    ws.auto_filter.ref = f"A1:I{rr-1}"

    # ============================================== 5. Contribuciones
    if contrib:
        ws = wb.create_sheet("Contribuciones")
        ws.column_dimensions['A'].width = 34
        for col in 'BCDEFG':
            ws.column_dimensions[col].width = 18
        ws.cell(row=1, column=1, value="Contribuciones patronales, Meta4 contra Axton").font = \
            Font(bold=True, size=13, color=AZUL)
        ws.cell(row=2, column=1, value=
                "Una contribución se calcula sobre la base, así que casi nunca es un error "
                "propio: la columna «De dónde viene» dice si hay que corregir el haber, la "
                "base o la contribución.").font = Font(size=9, italic=True, color="404040")

        cab(ws, [("Contribución", 34), ("Col. Meta4", 16), ("Cód. Axton", 14),
                 ("Total Meta4", 18), ("Total Axton", 18), ("Axton − Meta4", 18),
                 ("Legajos que difieren", 14)], fila=4)
        rr = 5
        for g in contrib['porConcepto']:
            vals = [g['concepto'], g['colM4'],
                    g['codAx'] if g['existeAx'] else f"{g['codAx']} (no hay columna)",
                    g['meta4'], g['axton'], g['dif'], g['legajos']]
            for i, val in enumerate(vals, 1):
                c = ws.cell(row=rr, column=i, value=val)
                c.font = N; c.border = BOX
                if i in (4, 5, 6):
                    c.number_format = MON
            if abs(g['dif']) > ruido:
                ws.cell(row=rr, column=6).font = Font(size=10, bold=True, color=ROJO)
            else:
                ws.cell(row=rr, column=7).font = Font(size=10, color=VERDE)
            rr += 1
        ws.cell(row=rr, column=1, value="TOTAL").font = B
        for i, k in ((4, 'meta4'), (5, 'axton'), (6, 'dif')):
            c = ws.cell(row=rr, column=i,
                        value=round(sum(g[k] for g in contrib['porConcepto']), 2))
            c.number_format = MON; c.font = B

        rr += 3
        ws.cell(row=rr, column=1, value="El detalle, legajo por legajo").font = \
            Font(bold=True, size=12, color=AZUL)
        rr += 1
        cab(ws, [("Legajo", 10), ("Apellido y Nombre", 30), ("Contribución", 30),
                 ("Meta4", 16), ("Axton", 16), ("Axton − Meta4", 16),
                 ("De dónde viene", 16), ("Por qué", 110)], fila=rr)
        rr += 1
        orden = {'Contribución': 0, 'Base': 1, 'Remuneración': 2, 'Redondeo': 3}
        for f in sorted(contrib['filas'],
                        key=lambda f: (orden.get(f['causa'], 9), -abs(f['dif']))):
            for i, val in enumerate([f['legajo'], f['nombre'], f['concepto'], f['meta4'],
                                     f['axton'], f['dif'], f['causa'], f['comentario']], 1):
                c = ws.cell(row=rr, column=i, value=val)
                c.font = N; c.border = BOX
                if i in (4, 5, 6):
                    c.number_format = MON
            color = {'Contribución': ROJO, 'Base': ROJO,
                     'Remuneración': AMBAR}.get(f['causa'], "404040")
            ws.cell(row=rr, column=7).font = Font(size=10, bold=f['causa'] != 'Redondeo',
                                                  color=color)
            ws.cell(row=rr, column=6).font = Font(size=10, bold=f['causa'] != 'Redondeo',
                                                  color=color)
            rr += 1
        if not contrib['filas']:
            ws.cell(row=rr, column=1, value="Ninguna: todas las contribuciones coinciden").font = \
                Font(size=10, color=VERDE)
            rr += 1

        if contrib['sinPar']:
            rr += 2
            ws.cell(row=rr, column=1,
                    value="Contribuciones de Axton que no están declaradas en el config").font = \
                Font(bold=True, size=12, color=AZUL)
            rr += 1
            cab(ws, [("Cód. Axton", 14), ("Concepto", 40), ("Importe del mes", 18)], fila=rr)
            rr += 1
            for c, nom, tot in contrib['sinPar']:
                for i, val in enumerate([c, nom, tot], 1):
                    cel = ws.cell(row=rr, column=i, value=val)
                    cel.font = N; cel.border = BOX
                    if i == 3:
                        cel.number_format = MON
                rr += 1

    # ============================================== 6. Sin comparar
    ws = wb.create_sheet("Sin comparar")
    for col, w in (('A', 28), ('B', 18), ('C', 46), ('D', 18), ('E', 66)):
        ws.column_dimensions[col].width = w
    st = {'r': 1}

    def bloque(t):
        ws.cell(row=st['r'], column=1, value=t).font = Font(bold=True, size=12, color=AZUL)
        st['r'] += 1

    def cabecera(*cs):
        for i, t in enumerate(cs, 1):
            c = ws.cell(row=st['r'], column=i, value=t)
            c.font = H; c.fill = HF
        st['r'] += 1

    def fila(*vals, money=()):
        for i, v in enumerate(vals, 1):
            c = ws.cell(row=st['r'], column=i, value=v)
            c.font = N
            if i in money:
                c.number_format = MON
        st['r'] += 1

    bloque("1. Conceptos de Meta4 con plata que en Axton no tienen columna")
    cabecera("Situación", "Cód. Meta4", "Concepto", "Importe en Meta4", "Qué significa")
    t = collections.defaultdict(float)
    for f in detalle:
        if f['estado'] == 'Falta en Axton (no hay columna)' and abs(f['meta4'] or 0) > tol:
            t[(f['codM4'], f['codAx'], f['concepto'], f['legajo'])] += f['meta4']
    if not t:
        fila("—", "", "Ninguno", None, "")
    for (cm, ca, nom, leg), v in sorted(t.items(), key=lambda kv: -abs(kv[1])):
        fila("No existe en Axton", cm, f"{nom} → esperado en Axton {ca}", round(v, 2),
             f"Legajo {leg}. El código de Axton de la tabla no es una columna del tabulado",
             money=(4,))
    st['r'] += 1

    bloque("2. Conceptos con plata que NO están en la tabla de equivalencias")
    cabecera("Sistema", "Código", "Concepto", "Importe", "Qué hice")
    em, ea = collections.defaultdict(float), collections.defaultdict(float)
    for v in porLegajo:
        for c, x in v['extraM4']:
            em[c] += x
        for c, x in v['extraAx']:
            ea[c] += x
    if not em and not ea:
        fila("—", "", "Ninguno", None, "")
    for c, v in sorted(em.items(), key=lambda kv: -abs(kv[1])):
        fila("Meta4", c, res['labels']['meta4'].get(c, ''), round(v, 2),
             "Sin equivalencia declarada: no se comparó, se informa", money=(4,))
    for c, v in sorted(ea.items(), key=lambda kv: -abs(kv[1])):
        fila("Axton", c, res['labels']['axton'].get(c, ''), round(v, 2),
             "Sin equivalencia declarada: no se comparó, se informa", money=(4,))
    st['r'] += 1

    bloque("3. Un concepto de Axton que vale por dos de Meta4 (se compara contra la suma)")
    cabecera("Cód. Axton", "Cód. Meta4", "Conceptos de Meta4 que se sumaron")
    hubo = False
    for g in sorted(res['grupos'], key=lambda g: g['axton']):
        if len(g['meta4']) < 2:
            continue
        hubo = True
        fila(g['axton'], " + ".join(g['meta4']), g['nombre'])
    if not hubo:
        fila("—", "", "Ninguno")
    st['r'] += 1

    bloque("4. Un concepto de Meta4 contra dos columnas de Axton (declarado a mano)")
    cabecera("Cód. Meta4", "Cód. Axton", "Concepto", "", "Qué hay que hacer")
    declarados = meta.get('mapeosDeclarados', {})
    if not declarados:
        fila("—", "", "Ninguno")
    for ax, extras in declarados.items():
        g = next((g for g in res['grupos'] if g['axton'] == ax), None)
        fila(" + ".join(g['meta4']) if g else '', ax + " + " + " + ".join(extras),
             g['nombre'] if g else '', "",
             "Axton liquida en una columna que la tabla no declara: confirmar y "
             "corregir la tabla de equivalencias")
    st['r'] += 1

    bloque("5. Renglones de la tabla de equivalencias que no se pudieron usar")
    cabecera("Fila del excel", "Cód. Axton", "Cód. Meta4", "Concepto", "Motivo")
    if not meta.get('sinAxton') and not meta.get('sinMeta4'):
        fila("—", "", "", "Ninguno", "")
    for p in meta.get('sinAxton', []):
        fila(p['fila'], p['axton'], p['meta4'], p['nombre'],
             "El código de Axton no es un número")
    for p in meta.get('sinMeta4', []):
        fila(p['fila'], p['axton'], p['meta4'], p['nombre'], "No trae código de Meta4")
    st['r'] += 1

    bloque("6. Columnas que quedaron fuera del cruce a propósito")
    cabecera("Sistema", "Códigos", "Qué son")
    cl = res['clasificacion']
    noConcepto = [c for c in res['labels']['meta4']
                  if c not in set(cl['habsM4']) and c not in set(cl['dtosM4'])]
    fila("Meta4", ", ".join(sorted(noConcepto, key=int))[:900],
         "Días, unidades, porcentajes y acumuladores: no son importes liquidados")
    fila("Axton", ", ".join(c for c in cl['basesAx'] if c not in cl['patronales'])[:900],
         "Bases y acumuladores de cálculo: no entran al neto (verificado con la "
         "aritmética del propio archivo)")
    fila("Axton", ", ".join(cl['patronales'])[:900],
         "ART y seguros: contribución patronal, no afecta el neto del empleado")

    wb.save(salida)
    return {'legajos': len(porLegajo), 'conDiferencia': len(conDifReal),
            'difTotal': round(totAx - totM4, 2),
            'contribuciones': len(contrib['filas']) if contrib else None}
