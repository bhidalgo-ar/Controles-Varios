# -*- coding: utf-8 -*-
"""Lectores de los archivos de un paralelo Meta4 -> Axton.

Cada lector devuelve estructura cruda, sin criterio de negocio: quien decide
que es haber y que es descuento es cruce.py con el config del cliente.

Regla que atraviesa todo: una celda vacia es None (no hay dato), nunca 0.
Un 0 es un dato que vale cero. Se pierden hallazgos si se confunden.
"""
import re
import collections

try:
    import openpyxl
    from lxml import html as LH
except ImportError as e:  # pragma: no cover
    raise SystemExit(
        f"Falta una libreria de Python: {e.name}.\n"
        "Instalar con:  pip install openpyxl lxml pymupdf"
    )


# ---------------------------------------------------------------- utilidades

def toNum(v):
    """Importe de una celda. None si no hay dato. Acepta 1.234,56 y 1234.56."""
    if v is None:
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace('\xa0', ' ').strip()
    if s in ('', '-', '&nbsp;', '#N/A', 'N/A'):
        return None
    s = s.replace(' ', '')
    neg = s.startswith('(') and s.endswith(')')
    if neg:
        s = s[1:-1]
    if ',' in s and '.' in s:
        s = s.replace('.', '').replace(',', '.')   # 1.234,56
    elif ',' in s:
        s = s.replace(',', '.')                    # 1234,56
    try:
        n = float(s)
    except ValueError:
        return None
    return -n if neg else n


def legajoKey(v, modo='sinCeros'):
    """Clave de legajo. Por default '007' y '7' son el MISMO empleado.

    Es la misma regla que `makeLegajoKey()` de la app (D-038/D-042). Nunca
    usar parseInt a mano: colapsa '12-B' y '12-C' en 12.
    """
    s = '' if v is None else str(v).strip()
    if s.endswith('.0'):
        s = s[:-2]
    if modo == 'literal':
        return s
    return str(int(s)) if re.fullmatch(r'\d+', s) else s.upper()


# ------------------------------------------------------- Meta4: tabulado_h

def leerMeta4(path, cfg, modoLegajo='sinCeros'):
    """Tabulado horizontal de Meta4 (.xlsx).

    Fila 1 encabezados; las columnas de concepto vienen como 'CODIGO-NOMBRE'.
    Trae UNA FILA POR LIQUIDACION: un legajo con la quincena y un ajuste del
    mismo mes aparece dos veces. No consolidar aca; lo hace cruce.py.
    """
    ws = openpyxl.load_workbook(path, data_only=True).active
    hdr = [('' if c.value is None else str(c.value).strip()) for c in ws[1]]

    cols, labels = {}, {}
    for i, h in enumerate(hdr):
        m = re.match(r'^(\d+)\s*-\s*(.*)$', h)
        if m:
            cols[m.group(1)] = i
            labels[m.group(1)] = m.group(2).strip()

    def idx(nombre, obligatoria=True):
        if nombre in hdr:
            return hdr.index(nombre)
        if obligatoria:
            raise SystemExit(
                f"El Tabulado de Meta4 no trae la columna '{nombre}'.\n"
                f"Encontre estos encabezados: {', '.join(h for h in hdr[:20] if h)}...\n"
                "Corregir el nombre en el config del cliente."
            )
        return None

    iLeg = idx(cfg['columnaLegajo'])
    iNom = idx(cfg.get('columnaNombre'), obligatoria=False) if cfg.get('columnaNombre') else None
    iNeto = idx(cfg['columnaNeto'])
    iDto = idx(cfg['columnaTotalDescuento'], obligatoria=False) if cfg.get('columnaTotalDescuento') else None

    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None or str(v).strip() == '' for v in r):
            continue
        leg = legajoKey(r[iLeg], modoLegajo)
        if not leg:
            continue
        rows.append({
            'legajo': leg,
            'nombre': '' if iNom is None else str(r[iNom] or '').strip(),
            'neto': toNum(r[iNeto]),
            'totalDescuento': None if iDto is None else toNum(r[iDto]),
            'conceptos': {c: toNum(r[i]) for c, i in cols.items()},
        })
    if not rows:
        raise SystemExit("El Tabulado de Meta4 no trajo ninguna fila de datos.")
    return {'rows': rows, 'labels': labels,
            'codigos': sorted(cols, key=lambda x: int(x)), 'headers': hdr}


# ------------------------------------------------------------- Axton: .xls

def leerAxton(path, modoLegajo='sinCeros'):
    """Tabulado de Axton ("Resumen de Liquidacion").

    Viene con extension .xls pero es HTML: una tabla, la fila TOTAL GENERAL
    ANTES de los encabezados, despues una fila de encabezados 'CODIGO - Nombre'
    y una subfila 'Imp'/'Cant'. Se ubica por firma, nunca por posicion.
    """
    raw = open(path, 'rb').read().decode('iso-8859-1', errors='replace')
    doc = LH.fromstring(raw)
    spans = doc.xpath('//span')
    preambulo = (spans[0].text_content() or '').strip() if spans else ''

    def celdas(tr):
        return [(c.text_content() or '').replace('\xa0', ' ').strip()
                for c in tr.xpath('./td|./th')]

    header, sub, total, data = None, None, None, []
    for tr in doc.xpath('//tr'):
        tags = {c.tag for c in tr.xpath('./td|./th')}
        cc = celdas(tr)
        if 'th' in tags:
            if header is None and any(c.startswith('Legajo') for c in cc):
                header = cc
            elif set(x for x in cc if x) <= {'Imp', 'Cant'}:
                sub = cc
            continue
        if cc and cc[0].upper().startswith('TOTAL GENERAL'):
            total = cc
            continue
        if cc and cc[0]:
            data.append(cc)

    if header is None:
        raise SystemExit(
            "No encontre la fila de encabezados del Tabulado de Axton "
            "(esperaba una fila con 'Legajo' y 'Apellido y Nombre').\n"
            "Puede ser que el archivo no sea el 'Resumen de Liquidacion'."
        )

    cols, labels = {}, {}
    for i, h in enumerate(header):
        m = re.match(r'^(\d+)\s*-\s*(.*)$', h)
        if m:
            cols[m.group(1)] = i
            labels[m.group(1)] = m.group(2).strip()
    fijas = {h: i for i, h in enumerate(header)}
    for req in ('Legajo', 'Neto'):
        if req not in fijas:
            raise SystemExit(f"El Tabulado de Axton no trae la columna '{req}'.")

    def val(cc, nombre):
        i = fijas.get(nombre)
        return None if i is None else toNum(cc[i])

    rows = []
    for cc in data:
        cc = cc + [''] * (len(header) - len(cc))
        rows.append({
            'legajo': legajoKey(cc[fijas['Legajo']], modoLegajo),
            'nombre': cc[fijas['Apellido y Nombre']] if 'Apellido y Nombre' in fijas else '',
            'neto': val(cc, 'Neto'),
            'bruto': val(cc, 'Bruto'),
            'retenciones': val(cc, 'Retenciones'),
            'exento': val(cc, 'Exento'),
            'liquidacion': cc[-1],
            'conceptos': {c: toNum(cc[i]) for c, i in cols.items()},
        })
    if not rows:
        raise SystemExit("El Tabulado de Axton no trajo ninguna fila de datos.")
    return {'rows': rows, 'labels': labels,
            'codigos': sorted(cols, key=lambda x: int(x)),
            'header': header, 'sub': sub, 'total': total,
            'preambulo': preambulo,
            'tipos': sorted({r['liquidacion'] for r in rows})}


# ------------------------------------------ PDF "Control de liquidacion"

def leerPdf(path, modoLegajo='sinCeros'):
    """PDF de la liqui de Meta4: un bloque por liquidacion + un total al final.

    Devuelve {legajo: {neto, haberes, descuentos, bloques}} y el total general.
    Es el ancla: si el Excel de Meta4 no reproduce esto, no se cruza nada.
    """
    try:
        import pymupdf
    except ImportError:
        try:
            import fitz as pymupdf
        except ImportError:
            raise SystemExit("Falta pymupdf. Instalar con:  pip install pymupdf")

    doc = pymupdf.open(path)
    lineas = [l.strip() for p in doc for l in p.get_text().split('\n')]
    arranques = [i for i, l in enumerate(lineas) if l == 'Empleado:']
    if not arranques:
        raise SystemExit(
            "El PDF no parece el 'Control de liquidacion' de Meta4: "
            "no encontre ningun bloque que arranque con 'Empleado:'."
        )

    CAMPOS = (('Total Netos:', 'neto'), ('Total Haberes:', 'haberes'),
              ('Total Descuentos:', 'descuentos'), ('Total Imponible:', 'imponible'))

    def campo(blk, etiqueta):
        for i, l in enumerate(blk):
            if l == etiqueta and i + 1 < len(blk):
                return toNum(blk[i + 1])
        return None

    porLegajo = collections.defaultdict(
        lambda: {'neto': 0.0, 'haberes': 0.0, 'descuentos': 0.0, 'bloques': 0})
    for k, i in enumerate(arranques):
        j = arranques[k + 1] if k + 1 < len(arranques) else len(lineas)
        blk = lineas[i:j]
        leg = None
        for t, l in enumerate(blk):
            if l == 'Legajo:' and t + 1 < len(blk):
                leg = legajoKey(blk[t + 1], modoLegajo)
                break
        if leg is None:
            continue
        d = porLegajo[leg]
        d['bloques'] += 1
        for etiqueta, clave in CAMPOS:
            v = campo(blk, etiqueta)
            if v is not None and clave in d:
                d[clave] += v

    # el bloque final (total general) no tiene 'Empleado:' propio: queda suelto
    # despues del ultimo bloque, asi que se lee del texto completo hacia atras.
    total = {}
    cola = lineas[arranques[-1]:]
    ultimo = None
    for i, l in enumerate(cola):
        if l == 'Total Netos:':
            ultimo = i
    if ultimo is not None:
        for etiqueta, clave in CAMPOS:
            for i in range(len(cola) - 1, 0, -1):
                if cola[i] == etiqueta and i + 1 < len(cola):
                    total[clave] = toNum(cola[i + 1])
                    break
    return {'legajos': dict(porLegajo), 'totalDeclarado': total,
            'bloques': len(arranques)}


# ------------------------------------------------- tabla de equivalencias

def leerEquivalencias(path):
    """Excel de 3 columnas: codigo Axton | codigo Meta4 | nombre del concepto.

    Los renglones donde alguno de los dos codigos no es un numero (#N/A, '*',
    un texto) NO se descartan en silencio: salen en 'sinAxton'/'sinMeta4' y el
    Excel los muestra, porque un concepto sin equivalencia es justo lo que un
    paralelo tiene que encontrar.
    """
    ws = openpyxl.load_workbook(path, data_only=True).active
    num = re.compile(r'^\d+$')
    pares, sinAx, sinM4 = [], [], []
    for i, r in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        ax = '' if len(r) < 1 or r[0] is None else str(r[0]).strip()
        m4 = '' if len(r) < 2 or r[1] is None else str(r[1]).strip()
        nom = '' if len(r) < 3 or r[2] is None else str(r[2]).strip()
        if not (ax or m4 or nom):
            continue
        rec = {'fila': i, 'axton': ax, 'meta4': m4, 'nombre': nom}
        if not num.match(ax):
            sinAx.append(rec)
        elif not num.match(m4):
            sinM4.append(rec)
        else:
            pares.append(rec)
    if not pares:
        raise SystemExit(
            "La tabla de equivalencias no trajo ningun par de codigos usable.\n"
            "Se esperan 3 columnas: codigo de Axton, codigo de Meta4, nombre."
        )
    porAxton = collections.defaultdict(list)
    for p in pares:
        porAxton[p['axton']].append(p)
    return {'pares': pares, 'sinAxton': sinAx, 'sinMeta4': sinM4,
            'porAxton': dict(porAxton),
            'porMeta4': {p['meta4']: p for p in pares}}


# ------------------------------- Meta4: control de cargas sociales

def leerCargas(path, cfg, modoLegajo='sinCeros'):
    """Export 'Control de cargas sociales' de Meta4 (.xlsx).

    Fila 1 encabezados y una columna por concepto de carga, con el NOMBRE del
    concepto y no su codigo: es otro export, no el Tabulado. Igual que el
    Tabulado trae UNA FILA POR LIQUIDACION, asi que no consolidar aca.

    Ademas de las contribuciones trae los aportes del empleado (TOT_JUB,
    TOT_LEY, TOT_OS). Esos no se cruzan contra Axton —ya viajan en el
    Tabulado— pero sirven de ancla: si no dan iguales a los del Tabulado, el
    archivo es de otra corrida y no se puede cruzar nada.
    """
    ws = openpyxl.load_workbook(path, data_only=True).active
    hdr = [('' if c.value is None else str(c.value).strip()) for c in ws[1]]

    def idx(nombre):
        if nombre not in hdr:
            raise SystemExit(
                f"El control de cargas sociales no trae la columna '{nombre}'.\n"
                f"Encontre estos encabezados: {', '.join(h for h in hdr if h)}\n"
                "Corregir el nombre en el bloque 'contribuciones' del config."
            )
        return hdr.index(nombre)

    iLeg = idx(cfg['columnaLegajo'])
    iNom = hdr.index(cfg['columnaNombre']) if cfg.get('columnaNombre') in hdr else None
    columnas = [h for h in hdr if h and h != cfg['columnaLegajo'] and h != cfg.get('columnaNombre')]

    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None or str(v).strip() == '' for v in r):
            continue
        leg = legajoKey(r[iLeg], modoLegajo)
        if not leg:
            continue
        rows.append({
            'legajo': leg,
            'nombre': '' if iNom is None else str(r[iNom] or '').strip(),
            'valores': {h: toNum(r[hdr.index(h)]) for h in columnas},
        })
    if not rows:
        raise SystemExit("El control de cargas sociales no trajo ninguna fila de datos.")
    return {'rows': rows, 'columnas': columnas, 'headers': hdr}
